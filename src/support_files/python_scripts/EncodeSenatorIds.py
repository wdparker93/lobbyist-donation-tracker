import openpyxl

path = 'C:\\Users\\wdpar\\vs_code_repos\\congressional-vote-tracker\\src\\support_files\\output_data_files\\PoliticianNameIdEncoder.xlsx'

recoderWB = openpyxl.load_workbook(path)
recoderSheet = recoderWB.active

fullNameEntryCol = 1
idEntryCol = 4
entryRow = 2

fullNameEntryCell = recoderSheet.cell(row = entryRow, column = fullNameEntryCol)
idEntryCell = recoderSheet.cell(row = entryRow, column = idEntryCol)

queryPrefix = 'UPDATE CONTRIBUTIONS_DATA SET RECIPIENT_ID = '
queryInfix = ' WHERE RECIPIENT_NAME = '
querySuffix = ';'
outputQueryArray = []

while fullNameEntryCell.value != "XXX":
    outputQuery = ''
    outputQuery += queryPrefix
    if fullNameEntryCell.value != "" and fullNameEntryCell.value != None:
        outputQuery += "'" + idEntryCell.value + "'"
        outputQuery += queryInfix
        outputQuery += "'" + fullNameEntryCell.value + "'"
        outputQuery += querySuffix
        outputQueryArray.append(outputQuery)
    entryRow += 1
    fullNameEntryCell = recoderSheet.cell(row = entryRow, column = fullNameEntryCol)
    idEntryCell = recoderSheet.cell(row = entryRow, column = idEntryCol)


outputFilePath = 'C:\\Users\\wdpar\\vs_code_repos\\congressional-vote-tracker\\src\\support_files\\output_query_txt_files\\senator_id_recoder_queries.txt'

with(open(outputFilePath, "a")) as f:
    for i in range(len(outputQueryArray)):
        print(outputQueryArray[i], file=f)
