import openpyxl

path = 'C:\\Users\\wdpar\\vs_code_repos\\congressional-vote-tracker\\src\\support_files\\Politician_Raw_Name_Recoder.xlsx'

recoderWB = openpyxl.load_workbook(path)
recoderSheet = recoderWB.active

entryCol = 1
entryRow = 2

delimitedList = ""

entryCell = recoderSheet.cell(row = entryRow, column = entryCol)
while entryCell.value != "XXX":
    if entryCell.value != "" and entryCell.value != None:
        delimitedList += "'" + entryCell.value + "', "
    entryRow += 1
    entryCell = recoderSheet.cell(row = entryRow, column = entryCol)

delimitedList = delimitedList[:-2]
outputCell = recoderSheet.cell(row = 2, column = 6)
outputCell.value = delimitedList

recoderWB.save('C:\\Users\\wdpar\\vs_code_repos\\congressional-vote-tracker\\src\\support_files\\Politician_Raw_Name_Recoder.xlsx')

#Save the output string to a text file so only one execution can be done in MySQL
excelFilePath = 'C:\\Users\\wdpar\\vs_code_repos\\congressional-vote-tracker\\src\\support_files\\Politician_Raw_Name_Recoder.xlsx'
outputFilePath = 'C:\\Users\\wdpar\\vs_code_repos\\congressional-vote-tracker\\src\\support_files\\recoder_query_file.txt'

updateQueryPrefix = "UPDATE contributions_data SET RECIPIENT_NAME = '"
updateQueryInfix = "' WHERE RECIPIENT_NAME_RAW IN ("
updateQuerySuffix = ');'

recoderWB = openpyxl.load_workbook(excelFilePath)
recoderSheet = recoderWB.active

targetNameCol = 3
targetNameRow = 2

outputListCol = 6
outputListRow = 2

targetName = ""
commaList = ""

outputQuery = ""

targetNameCell = recoderSheet.cell(row = targetNameRow, column = targetNameCol)
targetName = targetNameCell.value

outputListCell = recoderSheet.cell(row = outputListRow, column = outputListCol)
commaList = outputListCell.value

outputQuery += updateQueryPrefix
outputQuery += targetName
outputQuery += updateQueryInfix
outputQuery += commaList
outputQuery += updateQuerySuffix

with(open(outputFilePath, "a")) as f:
    print(outputQuery, file=f)
